# Python program to read an excel file
# import openpyxl module
import time
import openpyxl
import PathManage
import glob
import base64
import os;

#Avaliable Type :  Xint , int , string , Enum? , double , float

#print  (glob.glob('/Users/munjeonghwan/OneFolder/UnityProject_Launched/Utils_SimpleDataConverter/Xlsxs/*.xlsx'))
#print (os.getcwd())

start = time.time()  # 시작 시간 저장
xlsxPath = PathManage.xlsx_path + "/*.xlsx";

def makeBinaryByXlsx(wb_obj, table_name):
    allsheet = wb_obj.worksheets;
    string_a = ""
    for sheet_obj in allsheet:
        print(sheet_obj.title)
        if("Temp" in sheet_obj.title):
            continue
        string_a += sheet_obj.title + '\t\t'
        m_row = sheet_obj.max_row
        m_column = sheet_obj.max_column
        #1열은 변수 타입 , 2열은 변수명.
        for i in range(1, m_row + 1):
            if(sheet_obj.cell(row = i , column = 1).value is None):
                continue
            for j in range(1, m_column + 1):
                cell_obj = sheet_obj.cell(row = i, column = j)
                if(cell_obj.value is None):
                    continue
                if(i != 1 or j != 1):
                    string_a += '\t'
                string_a += str(cell_obj.value)
                #print(str(i) + "," + str(j) + " : " +  str(cell_obj.value))
                #print(string_a)
        string_a += '\t\t'
    utf8_encode = string_a.encode() 
    utf8_encode_base = base64.encodebytes(utf8_encode)
    file = open(PathManage.binary_path + "/" + table_name + ".bytes", "wb")
    file.write(utf8_encode_base)
    # file_proejct = open(PathManage.project_binary_path + "/" + table_name + ".bytes", "wb")
    # file_proejct.write(utf8_encode_base)
    

print("time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간