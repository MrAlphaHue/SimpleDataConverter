# Python program to read an excel file
# import openpyxl module
import time
import openpyxl
import glob
import base64
import os;

#Custom
import PathManage
import MakeBinaryFile
import MakeScriptFile
import MakeGeneralScriptFile

#Avaliable Type :  Xint , int , string , Enum? , double , float

#print  (glob.glob('/Users/munjeonghwan/OneFolder/UnityProject_Launched/Utils_SimpleDataConverter/Xlsxs/*.xlsx'))
#print  (os.getcwd())

start = time.time()  # 시작 시간 저장
xlsxPath = PathManage.xlsx_path + "/*.xlsx";

def CheckTabInCell(wb_obj, table_name):
    allsheet = wb_obj.worksheets;

    for sheet_obj in allsheet:
        if("TEMP" in sheet_obj.title.upper()):
            continue

        m_row = sheet_obj.max_row
        m_column = sheet_obj.max_column

        for i in range(1, m_row + 1):
            if(sheet_obj.cell(row = i , column = 1).value is None):
                continue
            for j in range(1, m_column + 1):
                cell_obj = sheet_obj.cell(row = i, column = j)
                if(cell_obj.value is None):
                    continue
                print("row " + str(i) + "column " + str(j) + " : " + str(cell_obj.value))
                if('\t' in str(cell_obj.value)):
                    print("Has Tab at " + table_name + " : [" + i + " ," + j +  "] !!!!!!!!")
                    return False
    return True
    

def func_getBasefileName(file_name):
    base_filename = os.path.splitext(file_name)
    base_filename = os.path.split(base_filename[0])
    base_filename = base_filename[1]
    return base_filename

##
MakeGeneralScriptFile.Generate()

##
for filename in glob.glob(xlsxPath):
    #with open(filename, 'r') as f:
    print(filename)
    if ("~$" in filename):
        continue

    wb_obj = openpyxl.load_workbook(filename, read_only=False, data_only=True) #read_only = true, really loosy.
    base_filename = func_getBasefileName(filename)
    if(CheckTabInCell(wb_obj, base_filename) is False):
        break
    MakeBinaryFile.makeBinaryByXlsx(wb_obj, base_filename)
    MakeScriptFile.makeScriptByXlsx(wb_obj, base_filename)

print("End time :", time.time() - start)  # 현재시각 -  시작시간 = 실행 시간